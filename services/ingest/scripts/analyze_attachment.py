import json
import os
import shutil
import sys
import zipfile
import re


POSITION_HINTS = ["职位", "岗位", "招考", "录用", "计划"]
SHEET_HINTS = ["职位", "岗位", "招考"]
MOJIBAKE_HINTS = "╔╗╚╝║═╟╢╤╧╨╩╪╫╬▓▒░■□∟└┘┌┐"


def sanitize_name(name: str) -> str:
    invalid = '<>:"/\\|?*'
    result = "".join("_" if char in invalid else char for char in name)
    return result[:160]


def looks_like_chinese(text: str) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", text or ""))


def looks_like_mojibake(text: str) -> bool:
    if not text:
        return False
    return any(char in MOJIBAKE_HINTS for char in text)


def repair_zip_name(name: str) -> str:
    if not name:
        return ""
    if looks_like_chinese(name):
        return name
    if not looks_like_mojibake(name):
        return name

    for encoding in ("gbk", "gb18030", "utf-8"):
        try:
            repaired = name.encode("cp437").decode(encoding)
        except Exception:
            continue
        if looks_like_chinese(repaired):
            return repaired
    return name


def score_name(name: str) -> int:
    score = 0
    lower = name.lower()
    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        score += 5
    if lower.endswith(".csv"):
        score += 3
    for hint in POSITION_HINTS:
        if hint in name:
            score += 10
    return score


def inspect_workbook(path: str):
    result = {
        "sheet_names": [],
        "preferred_sheets": [],
    }
    if not path.lower().endswith(".xlsx"):
        return result

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        result["sheet_names"] = list(workbook.sheetnames)
        result["preferred_sheets"] = [
            name for name in workbook.sheetnames if any(hint in name for hint in SHEET_HINTS)
        ]
        workbook.close()
    except Exception as exc:
        result["error"] = str(exc)
    return result


def extract_candidates(archive: zipfile.ZipFile, candidates, extract_dir):
    extracted = []
    if not extract_dir:
        return extracted

    os.makedirs(extract_dir, exist_ok=True)
    for candidate in candidates[:5]:
        archive_name = candidate["archive_name"]
        display_name = candidate["name"]
        target_name = sanitize_name(os.path.basename(display_name))
        target_path = os.path.join(extract_dir, target_name)

        with archive.open(archive_name) as source, open(target_path, "wb") as target:
            shutil.copyfileobj(source, target)

        extracted_item = {
            "name": display_name,
            "archive_name": archive_name,
            "path": target_path,
            "score": candidate["score"],
            "workbook": inspect_workbook(target_path),
        }
        extracted.append(extracted_item)
    return extracted


def analyze_file(path: str, extract_dir: str | None = None):
    payload = {
        "path": path,
        "kind": "file",
        "entries": [],
        "candidate_files": [],
        "extracted_files": [],
    }

    if zipfile.is_zipfile(path):
        payload["kind"] = "zip"
        with zipfile.ZipFile(path, "r") as archive:
            entries = []
            for info in archive.infolist():
                if info.is_dir():
                    continue
                repaired_name = repair_zip_name(info.filename)
                entry = {
                    "name": repaired_name,
                    "archive_name": info.filename,
                    "score": score_name(repaired_name),
                }
                entries.append(entry)
            payload["entries"] = sorted(entries, key=lambda item: item["score"], reverse=True)
            payload["candidate_files"] = [item for item in payload["entries"] if item["score"] > 0]
            payload["extracted_files"] = extract_candidates(archive, payload["candidate_files"], extract_dir)
        return payload

    payload["candidate_files"] = [
        {
            "name": os.path.basename(path),
            "score": score_name(os.path.basename(path)),
        }
    ]
    payload["workbook"] = inspect_workbook(path)
    return payload


def main():
    target = sys.argv[1]
    extract_dir = sys.argv[2] if len(sys.argv) > 2 else None
    result = analyze_file(target, extract_dir)
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    print(json.dumps(result, ensure_ascii=True))


if __name__ == "__main__":
    main()
