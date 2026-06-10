import json
import re
import sys
from typing import Any


HEADER_ALIASES = {
    "agency": ["招考单位", "用人单位", "部门名称", "招录机关"],
    "agencyCode": ["单位代码", "部门代码", "招录机关代码"],
    "title": ["招考职位", "职位名称", "职位", "招录职位"],
    "positionCode": ["职位代码", "招录职位代码", "岗位代码"],
    "description": ["职位简介", "职位描述", "工作内容"],
    "positionType": ["职位类型", "职位类别", "职务层次"],
    "headcount": ["录用人数", "招录人数", "招考人数", "人数"],
    "educationRaw": ["学历"],
    "degreeRaw": ["学位"],
    "majorPostgraduate": ["研究生专业名称及代码", "研究生专业", "研究生专业 名称及代码"],
    "majorUndergraduate": ["本科专业名称及代码", "本科专业", "本科专业 名称及代码"],
    "majorCollege": ["大专专业名称及代码", "专科专业名称及代码", "大专专业", "专科专业", "大专专业 名称及代码"],
    "serviceRequirement": ["是否要求2年以上基层工作经历", "基层工作经历", "基层工作最低年限"],
    "freshGraduateOnly": ["是否限应届毕业生报考", "应届毕业生", "限应届毕业生报考"],
    "politicalStatus": ["政治面貌"],
    "notes": ["其他要求", "其它要求", "备注"],
    "examArea": ["考区", "考试地点", "工作地点", "地区"],
}

CORE_FIELDS = {"agency", "title", "positionCode", "headcount"}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\u3000", " ")
    text = re.sub(r"\s+", "", text)
    return text.strip()


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_headcount(value: Any) -> int:
    text = clean_cell(value)
    match = re.search(r"\d+", text)
    return int(match.group(0)) if match else 0


def load_xls_rows(path: str):
    import xlrd

    workbook = xlrd.open_workbook(path)
    for sheet in workbook.sheets():
        rows = []
        for row_index in range(sheet.nrows):
            rows.append([sheet.cell_value(row_index, column) for column in range(sheet.ncols)])
        yield sheet.name, rows


def load_xlsx_rows(path: str):
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append(list(row))
            yield sheet_name, rows
    finally:
        workbook.close()


def iter_workbook_rows(path: str):
    lower = path.lower()
    if lower.endswith(".xls"):
      yield from load_xls_rows(path)
      return
    if lower.endswith(".xlsx"):
      yield from load_xlsx_rows(path)
      return
    raise ValueError(f"unsupported workbook format: {path}")


def resolve_field_map(header_cells):
    field_map = {}
    header_map = {}
    for index, raw_value in enumerate(header_cells):
        normalized = normalize_text(raw_value)
        if not normalized:
            continue
        header_map[index] = normalized
        for field, aliases in HEADER_ALIASES.items():
            if field in field_map:
                continue
            for alias in aliases:
                normalized_alias = normalize_text(alias)
                if normalized == normalized_alias or normalized_alias in normalized:
                    field_map[field] = index
                    break
    return field_map, header_map


def score_field_map(field_map):
    core_hits = sum(1 for field in CORE_FIELDS if field in field_map)
    optional_hits = len(field_map) - core_hits
    return core_hits * 20 + optional_hits * 3


def find_header(rows):
    best = None
    scan_limit = min(len(rows), 10)
    for row_index in range(scan_limit):
        field_map, header_map = resolve_field_map(rows[row_index])
        core_hits = sum(1 for field in CORE_FIELDS if field in field_map)
        if core_hits < 4:
            continue
        score = score_field_map(field_map)
        if not best or score > best["score"]:
            best = {
                "row_index": row_index,
                "field_map": field_map,
                "header_map": header_map,
                "score": score,
            }
    return best


def row_to_record(cells, field_map, sheet_name):
    record = {"sheetName": sheet_name}
    for field, column_index in field_map.items():
        if column_index >= len(cells):
            record[field] = ""
            continue
        record[field] = clean_cell(cells[column_index])
    record["headcountValue"] = parse_headcount(record.get("headcount", ""))
    if not record.get("examArea"):
        record["examArea"] = sheet_name
    return record


def is_position_record(record):
    if not record.get("agency") or not record.get("title"):
        return False
    if record.get("headcountValue", 0) <= 0:
        return False
    position_code = normalize_text(record.get("positionCode", ""))
    if not position_code:
        return False
    if "职位代码" in record.get("agency", "") or "职位代码" in record.get("title", ""):
        return False
    return True


def parse_sheet(sheet_name, rows):
    header = find_header(rows)
    if not header:
        return None

    records = []
    for row in rows[header["row_index"] + 1 :]:
        record = row_to_record(row, header["field_map"], sheet_name)
        if not is_position_record(record):
            continue
        records.append(record)

    if not records:
        return None

    return {
        "name": sheet_name,
        "headerRow": header["row_index"],
        "fieldMap": header["field_map"],
        "rowCount": len(records),
        "rows": records,
    }


def parse_workbook(path: str):
    sheet_results = []
    for sheet_name, rows in iter_workbook_rows(path):
        parsed = parse_sheet(sheet_name, rows)
        if parsed:
            sheet_results.append(parsed)

    total_rows = sum(sheet["rowCount"] for sheet in sheet_results)
    return {
        "path": path,
        "matched": total_rows > 0,
        "totalRows": total_rows,
        "sheets": [
            {
                "name": sheet["name"],
                "headerRow": sheet["headerRow"],
                "fieldMap": sheet["fieldMap"],
                "rowCount": sheet["rowCount"],
            }
            for sheet in sheet_results
        ],
        "rows": [row for sheet in sheet_results for row in sheet["rows"]],
    }


def choose_best_result(results):
    ranked = sorted(results, key=lambda item: item["totalRows"], reverse=True)
    return ranked[0] if ranked and ranked[0]["totalRows"] > 0 else None


def main():
    paths = sys.argv[1:]
    if not paths:
        raise SystemExit("missing workbook paths")

    results = []
    errors = []
    for path in paths:
        try:
            results.append(parse_workbook(path))
        except Exception as exc:
            errors.append({"path": path, "error": str(exc)})

    best = choose_best_result(results)
    payload = {
        "selected": best,
        "candidates": [
            {
                "path": item["path"],
                "matched": item["matched"],
                "totalRows": item["totalRows"],
                "sheetCount": len(item["sheets"]),
            }
            for item in results
        ],
        "errors": errors,
    }

    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    print(json.dumps(payload, ensure_ascii=True))


if __name__ == "__main__":
    main()
